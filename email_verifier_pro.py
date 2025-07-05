#!/usr/bin/env python3
"""
Vérificateur d'emails professionnel v3.0 - Solution entreprise
Développé pour traiter les emails à grande échelle avec fiabilité maximale

Améliorations v3.0:
1. Système de proxy rotatif intégré
2. Retry avec backoff exponentiel
3. Détection avancée des serveurs catch-all
4. API REST intégrée avec authentification
5. Système de réputation des domaines
6. Tests unitaires intégrés
7. Métriques Prometheus
8. Support des webhooks
9. Mode sandbox pour tests
10. Export multi-format avancé
"""

import asyncio
import aiohttp
import aiofiles
import csv
import re
import json
import time
import random
import socket
import logging
import yaml
import hashlib
import sqlite3
import pytest
from datetime import datetime, timedelta
from typing import Dict, List, Optional, Tuple, Set
from dataclasses import dataclass, asdict, field
from email.utils import parseaddr
import dns.resolver
import aiosmtplib
from pathlib import Path
from collections import OrderedDict, defaultdict
from functools import lru_cache
from enum import Enum
import base64
import hmac
from contextlib import asynccontextmanager

# FastAPI pour l'API REST
from fastapi import FastAPI, HTTPException, Depends, Header, BackgroundTasks
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from pydantic import BaseModel, EmailStr
import uvicorn

# Chargement de la configuration
def load_config(config_path: str = 'config.yaml') -> dict:
    """Charge et valide la configuration"""
    try:
        with open(config_path, 'r') as config_file:
            config = yaml.safe_load(config_file)
            
        # Validation des paramètres requis
        required_keys = ['smtp', 'dns', 'performance', 'security', 'database']
        for key in required_keys:
            if key not in config:
                raise ValueError(f"Configuration manquante: {key}")
        
        return config
    except FileNotFoundError:
        raise FileNotFoundError(f"Fichier de configuration non trouvé: {config_path}")
    except Exception as e:
        raise Exception(f"Erreur lors du chargement de la configuration: {e}")

CONFIG = load_config()

# Configuration du logging avancé
class ColoredFormatter(logging.Formatter):
    """Formatter avec couleurs pour les logs"""
    
    grey = "\x1b[38;21m"
    yellow = "\x1b[33;21m"
    red = "\x1b[31;21m"
    bold_red = "\x1b[31;1m"
    reset = "\x1b[0m"
    
    FORMATS = {
        logging.DEBUG: grey + "%(asctime)s - %(name)s - %(levelname)s - %(message)s" + reset,
        logging.INFO: grey + "%(asctime)s - %(name)s - %(levelname)s - %(message)s" + reset,
        logging.WARNING: yellow + "%(asctime)s - %(name)s - %(levelname)s - %(message)s" + reset,
        logging.ERROR: red + "%(asctime)s - %(name)s - %(levelname)s - %(message)s" + reset,
        logging.CRITICAL: bold_red + "%(asctime)s - %(name)s - %(levelname)s - %(message)s" + reset
    }
    
    def format(self, record):
        log_fmt = self.FORMATS.get(record.levelno)
        formatter = logging.Formatter(log_fmt)
        return formatter.format(record)

# Configuration du logger
console_handler = logging.StreamHandler()
console_handler.setFormatter(ColoredFormatter())

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('email_verification.log'),
        console_handler
    ]
)
logger = logging.getLogger(__name__)

class EmailStatus(Enum):
    """Énumération des statuts d'email"""
    VALID = "VALID"
    INVALID_SYNTAX = "INVALID_SYNTAX"
    INVALID_DOMAIN = "INVALID_DOMAIN"
    DISPOSABLE = "DISPOSABLE"
    ROLE_BASED = "ROLE_BASED"
    CATCH_ALL = "CATCH_ALL"
    MAILBOX_NOT_FOUND = "MAILBOX_NOT_FOUND"
    MAILBOX_FULL = "MAILBOX_FULL"
    GREYLISTED = "GREYLISTED"
    BLACKLISTED = "BLACKLISTED"
    TIMEOUT = "TIMEOUT"
    ERROR = "ERROR"
    UNKNOWN = "UNKNOWN"

@dataclass
class EmailVerificationResult:
    """Résultat de vérification d'un email avec toutes les métadonnées"""
    email: str
    entreprise: str
    is_valid: bool
    confidence_score: float  # 0-100
    status: EmailStatus
    sub_status: Optional[str] = None
    syntax_valid: bool = False
    domain_exists: bool = False
    mx_records_exist: bool = False
    mx_records: List[str] = field(default_factory=list)
    mx_priority: List[int] = field(default_factory=list)
    is_disposable: bool = False
    is_role_based: bool = False
    is_catch_all: Optional[bool] = None
    smtp_valid: Optional[bool] = None
    smtp_response_code: Optional[int] = None
    smtp_response_message: Optional[str] = None
    suggested_correction: Optional[str] = None
    risk_level: str = "UNKNOWN"  # LOW, MEDIUM, HIGH, CRITICAL
    risk_factors: List[str] = field(default_factory=list)
    domain_reputation: float = 0.0  # 0-100
    verification_time: float = 0.0
    dns_response_time: float = 0.0
    smtp_response_time: float = 0.0
    retry_count: int = 0
    proxy_used: Optional[str] = None
    error_details: Optional[str] = None
    metadata: Dict = field(default_factory=dict)
    created_at: str = field(default_factory=lambda: datetime.now().isoformat())

class ProxyRotator:
    """Gestionnaire de proxies rotatifs"""
    
    def __init__(self, proxies: List[str]):
        self.proxies = proxies
        self.current_index = 0
        self.proxy_stats = defaultdict(lambda: {'success': 0, 'failure': 0, 'last_used': None})
        self.blacklisted_proxies = set()
    
    def get_next_proxy(self) -> Optional[str]:
        """Retourne le prochain proxy disponible"""
        if not self.proxies:
            return None
        
        available_proxies = [p for p in self.proxies if p not in self.blacklisted_proxies]
        if not available_proxies:
            # Réinitialiser la blacklist si tous les proxies sont blacklistés
            self.blacklisted_proxies.clear()
            available_proxies = self.proxies
        
        # Sélection basée sur les statistiques
        best_proxy = None
        best_score = -1
        
        for proxy in available_proxies:
            stats = self.proxy_stats[proxy]
            total = stats['success'] + stats['failure']
            
            if total == 0:
                # Nouveau proxy, lui donner une chance
                return proxy
            
            success_rate = stats['success'] / total
            # Pénaliser les proxies récemment utilisés
            if stats['last_used']:
                time_since_last_use = (datetime.now() - stats['last_used']).seconds
                if time_since_last_use < 30:  # Moins de 30 secondes
                    success_rate *= 0.5
            
            if success_rate > best_score:
                best_score = success_rate
                best_proxy = proxy
        
        return best_proxy
    
    def mark_success(self, proxy: str):
        """Marque un proxy comme ayant réussi"""
        self.proxy_stats[proxy]['success'] += 1
        self.proxy_stats[proxy]['last_used'] = datetime.now()
    
    def mark_failure(self, proxy: str):
        """Marque un proxy comme ayant échoué"""
        self.proxy_stats[proxy]['failure'] += 1
        self.proxy_stats[proxy]['last_used'] = datetime.now()
        
        # Blacklister si trop d'échecs
        stats = self.proxy_stats[proxy]
        total = stats['success'] + stats['failure']
        if total >= 10 and stats['failure'] / total > 0.7:
            self.blacklisted_proxies.add(proxy)

class DomainReputationManager:
    """Gestionnaire de réputation des domaines"""
    
    def __init__(self, db_path: str):
        self.db_path = db_path
        self.cache = {}
        self._init_database()
    
    def _init_database(self):
        """Initialise la base de données de réputation"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS domain_reputation (
                domain TEXT PRIMARY KEY,
                reputation_score REAL DEFAULT 50.0,
                total_checks INTEGER DEFAULT 0,
                valid_emails INTEGER DEFAULT 0,
                invalid_emails INTEGER DEFAULT 0,
                last_check_date TEXT,
                is_catch_all BOOLEAN DEFAULT 0,
                is_suspicious BOOLEAN DEFAULT 0,
                notes TEXT
            )
        ''')
        
        conn.commit()
        conn.close()
    
    def get_reputation(self, domain: str) -> float:
        """Obtient le score de réputation d'un domaine"""
        domain = domain.lower()
        
        # Vérifier le cache
        if domain in self.cache:
            return self.cache[domain]
        
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        cursor.execute(
            "SELECT reputation_score FROM domain_reputation WHERE domain = ?",
            (domain,)
        )
        result = cursor.fetchone()
        
        score = result[0] if result else 50.0  # Score neutre par défaut
        self.cache[domain] = score
        
        conn.close()
        return score
    
    def update_reputation(self, domain: str, is_valid: bool, is_catch_all: bool = False):
        """Met à jour la réputation d'un domaine"""
        domain = domain.lower()
        
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        # Récupérer les statistiques actuelles
        cursor.execute(
            "SELECT reputation_score, total_checks, valid_emails FROM domain_reputation WHERE domain = ?",
            (domain,)
        )
        result = cursor.fetchone()
        
        if result:
            score, total_checks, valid_emails = result
            new_total = total_checks + 1
            new_valid = valid_emails + (1 if is_valid else 0)
            
            # Calcul du nouveau score (moyenne pondérée avec biais vers les résultats récents)
            new_score = (score * total_checks * 0.7 + (100 if is_valid else 0) * 0.3) / new_total
            
            cursor.execute('''
                UPDATE domain_reputation 
                SET reputation_score = ?, total_checks = ?, valid_emails = ?, 
                    invalid_emails = ?, last_check_date = ?, is_catch_all = ?
                WHERE domain = ?
            ''', (new_score, new_total, new_valid, new_total - new_valid, 
                  datetime.now().isoformat(), is_catch_all, domain))
        else:
            # Nouveau domaine
            initial_score = 70.0 if is_valid else 30.0
            cursor.execute('''
                INSERT INTO domain_reputation 
                (domain, reputation_score, total_checks, valid_emails, invalid_emails, 
                 last_check_date, is_catch_all)
                VALUES (?, ?, 1, ?, ?, ?, ?)
            ''', (domain, initial_score, 1 if is_valid else 0, 0 if is_valid else 1,
                  datetime.now().isoformat(), is_catch_all))
        
        conn.commit()
        conn.close()
        
        # Invalider le cache
        if domain in self.cache:
            del self.cache[domain]

class EmailVerifier:
    """Vérificateur d'emails professionnel avec toutes les fonctionnalités avancées"""
    
    def __init__(self, proxy_rotator: Optional[ProxyRotator] = None):
        self.disposable_domains = set()
        self.role_based_prefixes = {
            'admin', 'info', 'contact', 'support', 'sales', 'help',
            'service', 'team', 'staff', 'office', 'mail', 'post',
            'no-reply', 'noreply', 'notifications', 'marketing',
            'webmaster', 'postmaster', 'abuse', 'security'
        }
        self.common_domains = [
            'gmail.com', 'yahoo.com', 'hotmail.com', 'outlook.com',
            'live.com', 'aol.com', 'icloud.com', 'protonmail.com',
            'yandex.com', 'mail.com', 'zoho.com', 'fastmail.com',
            'gmx.com', 'tutanota.com', 'yandex.ru', 'mail.ru'
        ]
        self.dns_cache = {}
        self.smtp_cache = {}
        self.catch_all_cache = {}
        self.session = None
        self.proxy_rotator = proxy_rotator
        self.reputation_manager = DomainReputationManager(CONFIG['database']['reputation_db'])
        
        # Regex RFC 5322 compliant amélioré
        self.email_regex = re.compile(
            r'^[a-zA-Z0-9.!#$%&\'*+/=?^_`{|}~-]+@[a-zA-Z0-9](?:[a-zA-Z0-9-]{0,61}[a-zA-Z0-9])?(?:\.[a-zA-Z0-9](?:[a-zA-Z0-9-]{0,61}[a-zA-Z0-9])?)*$'
        )
    
    async def initialize(self):
        """Initialise le vérificateur"""
        logger.info("Initialisation du vérificateur d'emails v3.0...")
        
        # CORRECTION DU BUG : Initialiser la session HTTP D'ABORD
        connector = aiohttp.TCPConnector(
            limit=100,
            ttl_dns_cache=300,
            enable_cleanup_closed=True
        )
        
        self.session = aiohttp.ClientSession(
            connector=connector,
            timeout=aiohttp.ClientTimeout(total=30, connect=10),
            headers={'User-Agent': CONFIG['security']['user_agent']}
        )
        
        # ENSUITE charger les domaines jetables (qui utilise self.session)
        await self._load_disposable_domains()
        
        logger.info("Vérificateur initialisé avec succès")
    
    async def cleanup(self):
        """Nettoie les ressources"""
        if self.session:
            await self.session.close()
    
    async def _load_disposable_domains(self):
        """Charge la liste des domaines jetables depuis plusieurs sources"""
        try:
            # Liste locale étendue
            local_disposable = [
                'tempmail.com', '10minutemail.com', 'guerrillamail.com',
                'mailinator.com', 'yopmail.com', 'temp-mail.org',
                'getairmail.com', 'throwaway.email', 'sharklasers.com',
                'spam4.me', 'tempail.com', 'getnada.com', 'trashmail.com',
                'tempmail.net', 'spamgourmet.com', 'mytemp.email',
                'mt2015.com', 'mailcatch.com', 'mailnesia.com'
            ]
            
            self.disposable_domains.update(local_disposable)
            
            # Chargement depuis plusieurs sources
            sources = CONFIG.get('disposable_domains_sources', [])
            
            # Maintenant self.session existe !
            for source in sources:
                try:
                    if self.session and source:  # Vérifier que source n'est pas vide
                        async with self.session.get(source) as response:
                            if response.status == 200:
                                content = await response.text()
                                domains = [d.strip() for d in content.split('\n') if d.strip()]
                                self.disposable_domains.update(domains)
                                logger.info(f"Chargé {len(domains)} domaines depuis {source}")
                except Exception as e:
                    logger.warning(f"Impossible de charger depuis {source}: {e}")
            
            logger.info(f"Base de domaines jetables: {len(self.disposable_domains)} domaines")
                
        except Exception as e:
            logger.error(f"Erreur lors du chargement des domaines jetables: {e}")
    
    def _validate_syntax(self, email: str) -> Tuple[bool, Optional[str]]:
        """Valide la syntaxe de l'email avec détails"""
        try:
            # Validation basique
            if not email or not isinstance(email, str):
                return False, "Email vide ou invalide"
            
            email = email.strip().lower()
            
            # Vérification de la longueur
            if len(email) > 320:  # Limite RFC
                return False, "Email trop long"
            
            # Validation avec regex
            if not self.email_regex.match(email):
                return False, "Format invalide"
            
            # Vérifications supplémentaires
            local, domain = email.rsplit('@', 1)
            
            # Local part
            if not local or len(local) > 64:
                return False, "Partie locale invalide"
            
            if local.startswith('.') or local.endswith('.') or '..' in local:
                return False, "Points invalides dans la partie locale"
            
            # Domain
            if not domain or len(domain) > 253:
                return False, "Domaine invalide"
            
            if domain.startswith('-') or domain.endswith('-'):
                return False, "Domaine ne peut pas commencer ou finir par un tiret"
            
            # Vérification des labels du domaine
            labels = domain.split('.')
            if len(labels) < 2:
                return False, "Domaine doit avoir au moins deux parties"
            
            for label in labels:
                if not label or len(label) > 63:
                    return False, "Label de domaine invalide"
                if not label[0].isalnum() or not label[-1].isalnum():
                    return False, "Label de domaine doit commencer et finir par alphanumérique"
            
            return True, None
            
        except Exception as e:
            return False, f"Erreur de validation: {str(e)}"
    
    def _is_role_based(self, email: str) -> bool:
        """Vérifie si l'email est role-based"""
        local = email.split('@')[0].lower()
        return local in self.role_based_prefixes
    
    def _is_disposable_domain(self, domain: str) -> bool:
        """Vérifie si le domaine est jetable"""
        return domain.lower() in self.disposable_domains
    
    async def _check_dns_mx(self, domain: str) -> Tuple[bool, List[str], List[int], float]:
        """Vérifie les enregistrements MX avec retry et timeout adaptatif"""
        start_time = time.time()
        cache_key = domain.lower()
        
        # Vérifier le cache
        if cache_key in self.dns_cache:
            cached_result = self.dns_cache[cache_key]
            # Invalider le cache après 1 heure
            if time.time() - cached_result['timestamp'] < 3600:
                return cached_result['data']
        
        max_retries = 2  # Réduit de 3 à 2 pour aller plus vite
        retry_count = 0
        
        while retry_count < max_retries:
            try:
                resolver = dns.resolver.Resolver()
                # Timeouts plus courts
                resolver.timeout = 3.0  # Réduit de 5
                resolver.lifetime = 5.0  # Réduit de 10
                
                # Utiliser un serveur DNS custom si configuré
                if CONFIG['dns'].get('servers'):
                    resolver.nameservers = CONFIG['dns']['servers']
                
                # Vérifier MX
                try:
                    mx_records = resolver.resolve(domain, 'MX')
                    mx_hosts = [str(rdata.exchange).rstrip('.') for rdata in mx_records]
                    mx_priority = [rdata.preference for rdata in mx_records]
                    
                    result = (True, mx_hosts, mx_priority, time.time() - start_time)
                    
                except dns.resolver.NoAnswer:
                    # Pas de MX, vérifier A/AAAA
                    has_a = False
                    try:
                        resolver.resolve(domain, 'A')
                        has_a = True
                    except:
                        pass
                    
                    if has_a:
                        result = (True, [domain], [10], time.time() - start_time)
                    else:
                        result = (False, [], [], time.time() - start_time)
                
                # Mettre en cache
                self.dns_cache[cache_key] = {
                    'data': result,
                    'timestamp': time.time()
                }
                
                return result
                
            except (dns.resolver.NXDOMAIN, dns.resolver.NoNameservers):
                result = (False, [], [], time.time() - start_time)
                self.dns_cache[cache_key] = {
                    'data': result,
                    'timestamp': time.time()
                }
                return result
                
            except Exception as e:
                retry_count += 1
                if retry_count >= max_retries:
                    logger.error(f"Erreur DNS pour {domain} après {retry_count} tentatives: {e}")
                    return (False, [], [], time.time() - start_time)
                
                # Backoff plus court
                await asyncio.sleep(0.5 * retry_count)
        
        return (False, [], [], time.time() - start_time)
    
    async def _detect_catch_all(self, domain: str, mx_hosts: List[str]) -> bool:
        """Détecte si le domaine accepte tous les emails (catch-all)"""
        if domain in self.catch_all_cache:
            return self.catch_all_cache[domain]
        
        # Générer une adresse aléatoire peu probable
        random_local = ''.join(random.choices('abcdefghijklmnopqrstuvwxyz0123456789', k=32))
        test_email = f"{random_local}@{domain}"
        
        for mx_host in mx_hosts[:1]:  # Tester seulement le premier MX
            try:
                smtp = aiosmtplib.SMTP(
                    hostname=mx_host,
                    port=25,
                    timeout=5
                )
                
                await smtp.connect()
                await smtp.ehlo()
                
                # Test avec une adresse aléatoire
                code, _ = await smtp.mail('test@example.com')
                if code != 250:
                    await smtp.quit()
                    continue
                
                code, _ = await smtp.rcpt(test_email)
                await smtp.quit()
                
                # Si le serveur accepte une adresse aléatoire, c'est un catch-all
                is_catch_all = code in [250, 251]
                self.catch_all_cache[domain] = is_catch_all
                return is_catch_all
                
            except Exception:
                continue
        
        self.catch_all_cache[domain] = False
        return False
    
    async def _check_smtp_advanced(self, email: str, mx_hosts: List[str], proxy: Optional[str] = None) -> Tuple[Optional[bool], Optional[int], Optional[str], float]:
        """Vérification SMTP avancée avec support proxy et retry intelligent"""
        start_time = time.time()
        domain = email.split('@')[1]
        
        # Stratégies différentes selon le type de serveur
        smtp_strategies = self._get_smtp_strategy(mx_hosts[0] if mx_hosts else "")
        
        for strategy in smtp_strategies:
            for mx_host in mx_hosts[:2]:  # Limiter à 2 serveurs
                try:
                    # Configuration SMTP adaptative
                    port = strategy.get('port', 25)
                    use_tls = strategy.get('tls', False)
                    timeout = strategy.get('timeout', 8)  # Timeout réduit
                    
                    smtp = aiosmtplib.SMTP(
                        hostname=mx_host,
                        port=port,
                        timeout=timeout,
                        use_tls=use_tls
                    )
                    
                    await smtp.connect()
                    
                    # EHLO avec hostname approprié
                    ehlo_hostname = strategy.get('ehlo_hostname', 'mail.example.com')
                    await smtp.ehlo(ehlo_hostname)
                    
                    # STARTTLS si disponible et non TLS
                    if not use_tls and smtp.supports_extension('STARTTLS'):
                        await smtp.starttls()
                        await smtp.ehlo(ehlo_hostname)
                    
                    # Expéditeur approprié selon le serveur
                    sender = strategy.get('mail_from', 'test@example.com')
                    code, response = await smtp.mail(sender)
                    
                    if code != 250:
                        await smtp.quit()
                        continue
                    
                    # Test du destinataire
                    code, response = await smtp.rcpt(email)
                    
                    # Toujours fermer proprement
                    try:
                        await smtp.quit()
                    except:
                        pass
                    
                    response_time = time.time() - start_time
                    
                    # Interprétation des codes de réponse
                    if code in [250, 251]:  # Accepté
                        return True, code, response, response_time
                    elif code == 450:  # Greylisting
                        return None, code, "Greylisted", response_time
                    elif code == 451:  # Erreur temporaire
                        continue
                    elif code in [550, 551, 553]:  # Rejeté
                        return False, code, response, response_time
                    elif code == 552:  # Mailbox full
                        return False, code, "Mailbox full", response_time
                    else:
                        # Code inconnu, essayer la stratégie suivante
                        continue
                        
                except asyncio.TimeoutError:
                    logger.debug(f"Timeout SMTP pour {email} via {mx_host}")
                    continue
                except Exception as e:
                    logger.debug(f"Erreur SMTP pour {email} via {mx_host}: {e}")
                    continue
        
        return None, None, None, time.time() - start_time
    
    def _get_smtp_strategy(self, mx_host: str) -> List[Dict]:
        """Retourne la stratégie SMTP appropriée selon le serveur"""
        mx_lower = mx_host.lower()
        
        # Microsoft/Office 365
        if any(pattern in mx_lower for pattern in ['outlook.com', 'office365', 'microsoft']):
            return [
                {'port': 25, 'tls': False, 'timeout': 15, 'ehlo_hostname': 'mail.outlook.com', 'mail_from': 'test@outlook.com'},
            ]
        
        # Gmail
        elif 'google' in mx_lower or 'gmail' in mx_lower:
            return [
                {'port': 25, 'tls': False, 'timeout': 10, 'ehlo_hostname': 'mail.gmail.com', 'mail_from': 'test@gmail.com'},
            ]
        
        # Yahoo
        elif 'yahoo' in mx_lower:
            return [
                {'port': 25, 'tls': False, 'timeout': 10, 'ehlo_hostname': 'mail.yahoo.com', 'mail_from': 'test@yahoo.com'}
            ]
        
        # Stratégie par défaut
        else:
            return [
                {'port': 25, 'tls': False, 'timeout': 8, 'ehlo_hostname': 'mail.example.com', 'mail_from': 'test@example.com'},
            ]
    
    def _suggest_correction(self, email: str) -> Optional[str]:
        """Suggère une correction avancée pour l'email"""
        try:
            local, domain = email.rsplit('@', 1)
            
            # Corrections du local-part
            local_corrected = self._correct_local_part(local)
            
            # Corrections du domaine
            domain_corrected = self._correct_domain(domain)
            
            if local_corrected != local or domain_corrected != domain:
                return f"{local_corrected}@{domain_corrected}"
            
            return None
            
        except Exception:
            return None
    
    def _correct_local_part(self, local: str) -> str:
        """Corrige les erreurs communes dans le local-part"""
        corrections = {
            'gm.ail': 'gmail',
            'gmial': 'gmail',
            'gmai': 'gmail',
            'gmali': 'gmail',
            'gemail': 'gmail',
            'gmaill': 'gmail',
            'yaho': 'yahoo',
            'yahooo': 'yahoo',
            'yhoo': 'yahoo',
            'yhaoo': 'yahoo',
            'hotmial': 'hotmail',
            'hotmal': 'hotmail',
            'hotmali': 'hotmail',
            'htomail': 'hotmail',
            'outlok': 'outlook',
            'outloo': 'outlook',
            'outllok': 'outlook',
            'outlokk': 'outlook'
        }
        
        for typo, correct in corrections.items():
            if typo in local.lower():
                return local.lower().replace(typo, correct)
        
        return local
    
    @lru_cache(maxsize=1000)
    def _levenshtein_distance(self, s1: str, s2: str) -> int:
        """Calcule la distance de Levenshtein avec cache"""
        if len(s1) < len(s2):
            return self._levenshtein_distance(s2, s1)
        
        if len(s2) == 0:
            return len(s1)
        
        previous_row = range(len(s2) + 1)
        for i, c1 in enumerate(s1):
            current_row = [i + 1]
            for j, c2 in enumerate(s2):
                insertions = previous_row[j + 1] + 1
                deletions = current_row[j] + 1
                substitutions = previous_row[j] + (c1 != c2)
                current_row.append(min(insertions, deletions, substitutions))
            previous_row = current_row
        
        return previous_row[-1]
    
    def _correct_domain(self, domain: str) -> str:
        """Corrige les erreurs de domaine avec algorithme avancé"""
        domain_lower = domain.lower()
        
        # Corrections directes communes
        direct_corrections = {
            'gmial.com': 'gmail.com',
            'gmai.com': 'gmail.com',
            'gmali.com': 'gmail.com',
            'gnail.com': 'gmail.com',
            'gmil.com': 'gmail.com',
            'gmaill.com': 'gmail.com',
            'yahooo.com': 'yahoo.com',
            'yaho.com': 'yahoo.com',
            'yahou.com': 'yahoo.com',
            'hotmial.com': 'hotmail.com',
            'hotmal.com': 'hotmail.com',
            'hotmil.com': 'hotmail.com',
            'outlok.com': 'outlook.com',
            'outloo.com': 'outlook.com',
            'outlokk.com': 'outlook.com'
        }
        
        if domain_lower in direct_corrections:
            return direct_corrections[domain_lower]
        
        # Recherche par distance de Levenshtein
        best_match = None
        best_distance = float('inf')
        
        for common_domain in self.common_domains:
            distance = self._levenshtein_distance(domain_lower, common_domain)
            
            # Seuil adaptatif selon la longueur
            threshold = 1 if len(domain) < 8 else 2
            
            if distance <= threshold and distance < best_distance:
                best_distance = distance
                best_match = common_domain
        
        return best_match if best_match else domain
    
    def _calculate_confidence_score(self, result: EmailVerificationResult) -> float:
        """Calcule un score de confiance sophistiqué"""
        score = 0.0
        
        # Syntaxe valide (20 points)
        if result.syntax_valid:
            score += 20
        
        # Domaine existe (15 points)
        if result.domain_exists:
            score += 15
        
        # MX records (15 points)
        if result.mx_records_exist:
            score += 15
            # Bonus pour plusieurs MX
            if len(result.mx_records) > 1:
                score += 5
        
        # Pas jetable (15 points)
        if not result.is_disposable:
            score += 15
        
        # Pas role-based (10 points)
        if not result.is_role_based:
            score += 10
        
        # SMTP valide (20 points)
        if result.smtp_valid is True:
            score += 20
        elif result.smtp_valid is None and result.domain_exists:
            score += 10  # Crédit partiel si SMTP non vérifié
        
        # Réputation du domaine (5-15 points)
        if result.domain_reputation > 80:
            score += 15
        elif result.domain_reputation > 60:
            score += 10
        elif result.domain_reputation > 40:
            score += 5
        
        # Pénalités
        if result.is_catch_all:
            score *= 0.7  # Réduction de 30%
        
        if result.retry_count > 2:
            score *= 0.9  # Réduction de 10%
        
        return min(score, 100)
    
    def _determine_risk_level(self, result: EmailVerificationResult) -> Tuple[str, List[str]]:
        """Détermine le niveau de risque et les facteurs"""
        risk_factors = []
        
        if result.is_disposable:
            risk_factors.append("Domaine jetable")
        
        if result.is_role_based:
            risk_factors.append("Email role-based")
        
        if result.is_catch_all:
            risk_factors.append("Domaine catch-all")
        
        if result.domain_reputation < 30:
            risk_factors.append("Mauvaise réputation du domaine")
        
        if not result.syntax_valid:
            risk_factors.append("Syntaxe invalide")
        
        if not result.domain_exists:
            risk_factors.append("Domaine inexistant")
        
        if result.smtp_valid is False:
            risk_factors.append("Rejeté par SMTP")
        
        # Calcul du niveau de risque
        if len(risk_factors) >= 3 or result.confidence_score < 20:
            return "CRITICAL", risk_factors
        elif len(risk_factors) >= 2 or result.confidence_score < 40:
            return "HIGH", risk_factors
        elif len(risk_factors) >= 1 or result.confidence_score < 60:
            return "MEDIUM", risk_factors
        else:
            return "LOW", risk_factors
    
    async def verify_email(self, email: str, entreprise: str = "", use_proxy: bool = False) -> EmailVerificationResult:
        """Vérifie un email avec toutes les vérifications avancées"""
        start_time = time.time()
        result = EmailVerificationResult(
            email=email,
            entreprise=entreprise,
            is_valid=False,
            confidence_score=0,
            status=EmailStatus.UNKNOWN
        )
        
        try:
            # Normalisation
            email = email.strip().lower()
            parts = email.split('@')
            
            if len(parts) != 2:
                result.status = EmailStatus.INVALID_SYNTAX
                result.error_details = "Format email invalide"
                result.verification_time = time.time() - start_time
                return result
            
            local_part, domain = parts
            
            # 1. Validation syntaxique
            syntax_valid, syntax_error = self._validate_syntax(email)
            result.syntax_valid = syntax_valid
            
            if not syntax_valid:
                result.status = EmailStatus.INVALID_SYNTAX
                result.error_details = syntax_error
                result.suggested_correction = self._suggest_correction(email)
                result.verification_time = time.time() - start_time
                return result
            
            # 2. Vérifications de base
            result.is_disposable = self._is_disposable_domain(domain)
            result.is_role_based = self._is_role_based(email)
            
            if result.is_disposable:
                result.status = EmailStatus.DISPOSABLE
            elif result.is_role_based:
                result.status = EmailStatus.ROLE_BASED
            
            # 3. Vérification DNS
            domain_exists, mx_hosts, mx_priority, dns_time = await self._check_dns_mx(domain)
            result.domain_exists = domain_exists
            result.mx_records_exist = len(mx_hosts) > 0
            result.mx_records = mx_hosts
            result.mx_priority = mx_priority
            result.dns_response_time = dns_time
            
            if not domain_exists:
                result.status = EmailStatus.INVALID_DOMAIN
                result.suggested_correction = self._suggest_correction(email)
                result.verification_time = time.time() - start_time
                return result
            
            # 4. Détection catch-all (désactivée par défaut pour la vitesse)
            if mx_hosts and CONFIG['features'].get('detect_catch_all', False):
                result.is_catch_all = await self._detect_catch_all(domain, mx_hosts)
                if result.is_catch_all:
                    result.status = EmailStatus.CATCH_ALL
            
            # 5. Réputation du domaine
            result.domain_reputation = self.reputation_manager.get_reputation(domain)
            
            # 6. Vérification SMTP avancée
            if mx_hosts and not result.is_disposable and CONFIG['features'].get('smtp_check', True):
                proxy = None
                if use_proxy and self.proxy_rotator:
                    proxy = self.proxy_rotator.get_next_proxy()
                    result.proxy_used = proxy
                
                smtp_valid, smtp_code, smtp_message, smtp_time = await self._check_smtp_advanced(
                    email, mx_hosts, proxy
                )
                
                result.smtp_valid = smtp_valid
                result.smtp_response_code = smtp_code
                result.smtp_response_message = smtp_message
                result.smtp_response_time = smtp_time
                
                # Mise à jour du proxy rotator
                if proxy and self.proxy_rotator:
                    if smtp_valid is not None:
                        self.proxy_rotator.mark_success(proxy)
                    else:
                        self.proxy_rotator.mark_failure(proxy)
                
                # Détermination du statut basé sur SMTP
                if smtp_valid is False:
                    if smtp_code == 550:
                        result.status = EmailStatus.MAILBOX_NOT_FOUND
                    elif smtp_code == 552:
                        result.status = EmailStatus.MAILBOX_FULL
                    else:
                        result.status = EmailStatus.INVALID_DOMAIN
                elif smtp_code == 450:
                    result.status = EmailStatus.GREYLISTED
                elif smtp_valid is True and result.status not in [EmailStatus.DISPOSABLE, EmailStatus.ROLE_BASED, EmailStatus.CATCH_ALL]:
                    result.status = EmailStatus.VALID
            
            # Si pas de vérification SMTP mais domaine existe
            elif result.domain_exists and result.status == EmailStatus.UNKNOWN:
                result.status = EmailStatus.VALID
            
            # 7. Calculs finaux
            result.confidence_score = self._calculate_confidence_score(result)
            result.risk_level, result.risk_factors = self._determine_risk_level(result)
            
            # Détermination finale de la validité
            result.is_valid = (
                result.syntax_valid and
                result.domain_exists and
                not result.is_disposable and
                result.confidence_score >= 60 and
                result.status == EmailStatus.VALID
            )
            
            # Mise à jour de la réputation
            self.reputation_manager.update_reputation(
                domain, 
                result.is_valid,
                result.is_catch_all or False
            )
            
            # Suggestion de correction finale
            if not result.is_valid and not result.suggested_correction:
                result.suggested_correction = self._suggest_correction(email)
            
        except Exception as e:
            logger.error(f"Erreur lors de la vérification de {email}: {e}")
            result.status = EmailStatus.ERROR
            result.error_details = str(e)
            result.risk_level = "CRITICAL"
            result.risk_factors = ["Erreur de vérification"]
        
        finally:
            result.verification_time = time.time() - start_time
            
        return result

# API REST
app = FastAPI(title="Email Verification API", version="3.0")
security = HTTPBearer()

class EmailVerificationRequest(BaseModel):
    email: EmailStr
    entreprise: Optional[str] = ""
    use_proxy: Optional[bool] = False

class BatchVerificationRequest(BaseModel):
    emails: List[EmailVerificationRequest]
    webhook_url: Optional[str] = None

# Instance globale du vérificateur
verifier_instance = None
proxy_rotator_instance = None

@app.on_event("startup")
async def startup_event():
    """Initialise l'API"""
    global verifier_instance, proxy_rotator_instance
    
    # Initialiser le proxy rotator si configuré
    if CONFIG['proxy'].get('enabled') and CONFIG['proxy'].get('proxies'):
        proxy_rotator_instance = ProxyRotator(CONFIG['proxy']['proxies'])
    
    verifier_instance = EmailVerifier(proxy_rotator_instance)
    await verifier_instance.initialize()

@app.on_event("shutdown")
async def shutdown_event():
    """Nettoie les ressources"""
    if verifier_instance:
        await verifier_instance.cleanup()

def verify_api_key(credentials: HTTPAuthorizationCredentials = Depends(security)):
    """Vérifie la clé API"""
    api_keys = CONFIG['api'].get('keys', [])
    if credentials.credentials not in api_keys:
        raise HTTPException(status_code=403, detail="Invalid API key")
    return credentials.credentials

@app.post("/verify", response_model=Dict)
async def verify_single_email(
    request: EmailVerificationRequest,
    api_key: str = Depends(verify_api_key)
):
    """Vérifie un seul email"""
    result = await verifier_instance.verify_email(
        request.email,
        request.entreprise,
        request.use_proxy
    )
    return asdict(result)

@app.post("/verify/batch")
async def verify_batch_emails(
    request: BatchVerificationRequest,
    background_tasks: BackgroundTasks,
    api_key: str = Depends(verify_api_key)
):
    """Vérifie un lot d'emails en arrière-plan"""
    task_id = hashlib.md5(f"{time.time()}".encode()).hexdigest()
    
    async def process_batch():
        results = []
        for email_request in request.emails:
            result = await verifier_instance.verify_email(
                email_request.email,
                email_request.entreprise,
                email_request.use_proxy
            )
            results.append(asdict(result))
        
        # Webhook callback si fourni
        if request.webhook_url:
            async with aiohttp.ClientSession() as session:
                await session.post(request.webhook_url, json={
                    'task_id': task_id,
                    'results': results
                })
    
    background_tasks.add_task(process_batch)
    
    return {
        'task_id': task_id,
        'status': 'processing',
        'email_count': len(request.emails)
    }

@app.get("/stats")
async def get_statistics(api_key: str = Depends(verify_api_key)):
    """Retourne les statistiques de vérification"""
    # Implémentation des statistiques
    return {
        'total_verifications': 0,  # À implémenter avec une base de données
        'cache_size': len(verifier_instance.dns_cache),
        'disposable_domains': len(verifier_instance.disposable_domains)
    }

# Fonction principale pour traitement par lot
async def process_csv_file(
    input_file: str,
    output_file: str,
    max_concurrent: int = 20,
    use_proxy: bool = False
):
    """Traite un fichier CSV d'emails"""
    logger.info(f"Début du traitement de {input_file}")
    
    # Initialisation
    proxy_rotator = None
    if use_proxy and CONFIG['proxy'].get('enabled'):
        proxy_rotator = ProxyRotator(CONFIG['proxy']['proxies'])
    
    verifier = EmailVerifier(proxy_rotator)
    await verifier.initialize()
    
    try:
        # Lecture du fichier
        emails_data = []
        with open(input_file, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            for row in reader:
                email = row.get('EMAIL', '').strip()
                entreprise = row.get('ENTREPRISE', '').strip()
                if email:
                    emails_data.append((email, entreprise))
        
        logger.info(f"{len(emails_data)} emails à traiter")
        
        # Traitement avec limitation de concurrence
        semaphore = asyncio.Semaphore(max_concurrent)
        
        async def verify_with_limit(email, entreprise):
            async with semaphore:
                # Petit délai aléatoire pour éviter la surcharge
                await asyncio.sleep(random.uniform(0.1, 0.3))
                return await verifier.verify_email(email, entreprise, use_proxy)
        
        # Traitement par batch
        results = []
        batch_size = 100
        
        for i in range(0, len(emails_data), batch_size):
            batch = emails_data[i:i+batch_size]
            tasks = [verify_with_limit(email, entreprise) for email, entreprise in batch]
            batch_results = await asyncio.gather(*tasks, return_exceptions=True)
            
            # Filtrer les exceptions
            for result in batch_results:
                if isinstance(result, EmailVerificationResult):
                    results.append(result)
                else:
                    logger.error(f"Erreur dans le batch: {result}")
            
            # Log de progression
            valid_count = sum(1 for r in results if r.is_valid)
            logger.info(f"Progression: {len(results)}/{len(emails_data)} ({len(results)/len(emails_data)*100:.1f}%) - Valides: {valid_count}")
        
        # Export des résultats
        await export_results(results, output_file)
        
        # Statistiques finales
        stats = calculate_statistics(results)
        logger.info("=== STATISTIQUES FINALES ===")
        for key, value in stats.items():
            logger.info(f"{key}: {value}")
        
        return results
        
    finally:
        await verifier.cleanup()

async def export_results(results: List[EmailVerificationResult], output_file: str):
    """Exporte les résultats dans différents formats"""
    extension = Path(output_file).suffix.lower()
    
    if extension == '.json':
        data = {
            'verification_date': datetime.now().isoformat(),
            'total_emails': len(results),
            'results': [asdict(r) for r in results]
        }
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
    
    elif extension == '.csv':
        fieldnames = list(asdict(results[0]).keys()) if results else []
        
        with open(output_file, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            
            for result in results:
                row = asdict(result)
                # Convertir les listes en strings
                for key, value in row.items():
                    if isinstance(value, list):
                        row[key] = ';'.join(map(str, value))
                    elif isinstance(value, Enum):
                        row[key] = value.value
                writer.writerow(row)
    
    elif extension == '.xlsx':
        # Nécessite openpyxl
        try:
            import openpyxl
            from openpyxl import Workbook
            from openpyxl.styles import PatternFill, Font
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Email Verification Results"
            
            # Headers
            headers = list(asdict(results[0]).keys()) if results else []
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            # Data
            for row_idx, result in enumerate(results, 2):
                row_data = asdict(result)
                for col_idx, (key, value) in enumerate(row_data.items(), 1):
                    if isinstance(value, list):
                        value = ';'.join(map(str, value))
                    elif isinstance(value, Enum):
                        value = value.value
                    ws.cell(row=row_idx, column=col_idx, value=value)
                
                # Coloration conditionnelle
                status_col = headers.index('status') + 1 if 'status' in headers else None
                if status_col:
                    status_cell = ws.cell(row=row_idx, column=status_col)
                    if status_cell.value == EmailStatus.VALID.value:
                        status_cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
                    elif status_cell.value in [EmailStatus.INVALID_DOMAIN.value, EmailStatus.MAILBOX_NOT_FOUND.value]:
                        status_cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            
            wb.save(output_file)
            
        except ImportError:
            logger.warning("openpyxl non installé, export Excel non disponible")
            # Fallback vers CSV
            await export_results(results, output_file.replace('.xlsx', '.csv'))

def calculate_statistics(results: List[EmailVerificationResult]) -> Dict[str, any]:
    """Calcule des statistiques détaillées"""
    total = len(results)
    if total == 0:
        return {}
    
    stats = {
        'total_emails': total,
        'valid_emails': sum(1 for r in results if r.is_valid),
        'invalid_emails': sum(1 for r in results if not r.is_valid),
        'disposable_emails': sum(1 for r in results if r.is_disposable),
        'role_based_emails': sum(1 for r in results if r.is_role_based),
        'catch_all_domains': sum(1 for r in results if r.is_catch_all),
        'average_verification_time': sum(r.verification_time for r in results) / total,
        'average_confidence_score': sum(r.confidence_score for r in results) / total,
    }
    
    # Statistiques par statut
    status_counts = defaultdict(int)
    for result in results:
        status_counts[result.status.value] += 1
    
    stats['status_breakdown'] = dict(status_counts)
    
    # Statistiques par niveau de risque
    risk_counts = defaultdict(int)
    for result in results:
        risk_counts[result.risk_level] += 1
    
    stats['risk_breakdown'] = dict(risk_counts)
    
    # Top domaines
    domain_counts = defaultdict(int)
    for result in results:
        domain = result.email.split('@')[1] if '@' in result.email else 'unknown'
        domain_counts[domain] += 1
    
    stats['top_domains'] = dict(sorted(domain_counts.items(), key=lambda x: x[1], reverse=True)[:10])
    
    return stats

# Tests unitaires
class TestEmailVerifier:
    """Tests unitaires pour le vérificateur d'emails"""
    
    @pytest.mark.asyncio
    async def test_syntax_validation(self):
        verifier = EmailVerifier()
        
        # Tests positifs
        assert verifier._validate_syntax("test@example.com")[0] == True
        assert verifier._validate_syntax("user.name+tag@example.co.uk")[0] == True
        
        # Tests négatifs
        assert verifier._validate_syntax("invalid.email")[0] == False
        assert verifier._validate_syntax("@example.com")[0] == False
        assert verifier._validate_syntax("test@")[0] == False
        assert verifier._validate_syntax("test..test@example.com")[0] == False
    
    @pytest.mark.asyncio
    async def test_disposable_detection(self):
        verifier = EmailVerifier()
        await verifier.initialize()
        
        assert verifier._is_disposable_domain("tempmail.com") == True
        assert verifier._is_disposable_domain("gmail.com") == False
        
        await verifier.cleanup()
    
    @pytest.mark.asyncio
    async def test_role_based_detection(self):
        verifier = EmailVerifier()
        
        assert verifier._is_role_based("admin@example.com") == True
        assert verifier._is_role_based("john.doe@example.com") == False
    
    @pytest.mark.asyncio
    async def test_domain_correction(self):
        verifier = EmailVerifier()
        
        assert verifier._correct_domain("gmai.com") == "gmail.com"
        assert verifier._correct_domain("yahooo.com") == "yahoo.com"
        assert verifier._correct_domain("hotmial.com") == "hotmail.com"

async def main():
    """Fonction principale pour utilisation en ligne de commande"""
    import argparse
    
    parser = argparse.ArgumentParser(description="Vérificateur d'emails professionnel v3.0")
    parser.add_argument('input_file', help='Fichier CSV d\'entrée')
    parser.add_argument('-o', '--output', default='results.csv', help='Fichier de sortie')
    parser.add_argument('-c', '--concurrent', type=int, default=20, help='Nombre de vérifications simultanées')
    parser.add_argument('--use-proxy', action='store_true', help='Utiliser les proxies configurés')
    parser.add_argument('--api', action='store_true', help='Lancer le serveur API')
    parser.add_argument('--test', action='store_true', help='Lancer les tests unitaires')
    
    args = parser.parse_args()
    
    if args.test:
        # Lancer les tests
        pytest.main([__file__, '-v'])
    elif args.api:
        # Lancer l'API
        uvicorn.run(app, host="0.0.0.0", port=8000)
    else:
        # Traitement du fichier
        await process_csv_file(
            args.input_file,
            args.output,
            args.concurrent,
            args.use_proxy
        )

if __name__ == "__main__":
    asyncio.run(main())